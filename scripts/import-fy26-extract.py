#!/usr/bin/env python3
"""
Extract + categorize MASTER FY26.xlsx into a clean import.json for the HCEDP
tracker. Pure data transformation — no DB access. The JSON it emits is the
reviewable audit trail: every project row shows exactly what it became.

Scope (per Daniel's decisions):
  - FY26 Leads (107)        -> Projects, stage from Status ("Active"->RFI_SUBMITTED)
  - No Submission (23)      -> Projects, stage NO_SUBMISSION + reason
  - FY25 Follow-Ups (9)     -> Projects, archived
  - Site Tours (17)         -> SiteVisits, matched to a project by codename
Community tabs are NOT imported as projects (they duplicate the master); the
offered-site names are captured per project as a "Sites Offered" note for Phase 2.
"""
import openpyxl, re, json, sys, os
from datetime import datetime, date
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from fy26_util_parser import parse_utility  # normalizes the compact utility strings

#   python scripts/import-fy26-extract.py "<path to MASTER FY26.xlsx>" <out.json>
SRC = sys.argv[1] if len(sys.argv) > 1 else "C:/Users/Daniel/Downloads/MASTER FY26.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "import.json"

# --- header handling -------------------------------------------------------
def hnorm(h):
    if h is None: return ""
    return re.sub(r"\s+", " ", str(h).replace("\n", " ")).strip().lower()

# canonical field -> list of accepted header spellings (normalized)
HEADERS = {
    "name":        ["project name"],
    "active":      ["active date"],
    "submitted":   ["submission date"],
    "status":      ["status"],
    "proj_type":   ["project type"],
    "mfg_type":    ["manufacturing type"],
    "industry":    ["industry"],
    "sf_min":      ["building sf (min)"],
    "sf_max":      ["building sf (max)"],
    "ac_min":      ["acres (min)"],
    "ac_max":      ["acres (max)"],
    "jobs":        ["jobs"],
    "rail":        ["rail?", "rail"],
    "port":        ["port?", "port"],
    "existing":    ["existing building required?", "existing building?", "existing building"],
    "investment":  ["investment"],
    "source":      ["source"],
    "contact":     ["contact"],
    "power":       ["power"],
    "water":       ["water"],
    "sewer":       ["wastewater / sewer", "wastewater/sewer", "wastewater", "sewer"],
    "gas":         ["gas"],
    "bluebonnet":  ["bluebonnet sites", "bluebonnet"],
    "pedernales":  ["pedernales sites", "pedernales"],
    "origin":      ["international/ coming from?", "international/coming from?",
                    "international / coming from?", "international coming from?"],
    "competing":   ["competing states for project", "competing states"],
    "notes":       ["notes"],
    "updates":     ["updates"],
    "consultant":  ["site consultant(s) or company | (if known)",
                    "site consultant(s) or company (if known)",
                    "site consultant(s) or company"],
    "no_sub_why":  ["why did we not submit?"],
}

def build_index(ws):
    hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    norm_to_idx = {}
    for i, h in enumerate(hdr):
        n = hnorm(h)
        if n and n not in norm_to_idx:
            norm_to_idx[n] = i
    field_idx = {}
    for field, spellings in HEADERS.items():
        for sp in spellings:
            if sp in norm_to_idx:
                field_idx[field] = norm_to_idx[sp]
                break
    return field_idx

def get(row, idx, field):
    i = idx.get(field)
    if i is None or i >= len(row): return None
    v = row[i]
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v

# --- value coercion --------------------------------------------------------
def as_text(v):
    if v is None: return None
    if isinstance(v, (datetime, date)): return v.isoformat()
    s = str(v).strip()
    return s if s else None

def as_num(v):
    if v is None: return None
    if isinstance(v, (int, float)): return float(v)
    s = re.sub(r"[,$\s]", "", str(v))
    m = re.search(r"-?\d+(\.\d+)?", s)
    return float(m.group(0)) if m else None

def as_int(v):
    n = as_num(v)
    return int(round(n)) if n is not None else None

def as_date(v):
    if v is None: return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError: pass
    return None

EMPTY_UTIL = {"", "-", "–", "n/a", "na"}
def util_raw(v):
    t = as_text(v)
    if t is None: return None
    return None if t.lower() in EMPTY_UTIL else t

# --- enum mappings ---------------------------------------------------------
STATUS_TO_STAGE = {
    "active": "RFI_SUBMITTED",
    "short listed": "SHORTLISTED", "shortlisted": "SHORTLISTED",
    "negotiations": "IN_NEGOTIATIONS", "in negotiations": "IN_NEGOTIATIONS",
    "won": "WON", "lost": "LOST",
    "expansion": "RFI_SUBMITTED",   # data-entry glitch: really a project type
}
SOURCE_TO_ENUM = {
    "oa": "OPPORTUNITY_AUSTIN",
    "gov": "TEXAS_GOVERNORS_OFFICE", "governor": "TEXAS_GOVERNORS_OFFICE",
    "direct - company": "DIRECT_COMPANY",
    "direct - regional partner": "DIRECT_REGIONAL_PARTNERS",
    "direct - regional partners": "DIRECT_REGIONAL_PARTNERS",
    "direct - site selector": "DIRECT_SITE_SELECTOR",
    "direct - broker": "DIRECT_BROKER",
    "direct - marketing trip": "DIRECT_MARKETING_TRIP",
    "direct - other": "DIRECT_OTHER",
}
def pref(v):
    t = (as_text(v) or "").lower()
    if t.startswith("yes"): return "YES"
    if t.startswith("no"):  return "NO"
    if t.startswith("pref"): return "PREFERRED"
    return None  # e.g. "300 Miles" is not a tri-state answer

def stage_from_status(status_text, has_submit):
    if status_text is None:
        return "RFI_SUBMITTED" if has_submit else "RFI_RECEIVED"
    return STATUS_TO_STAGE.get(status_text.strip().lower(),
                               "RFI_SUBMITTED" if has_submit else "RFI_RECEIVED")

def lead_source(v):
    t = (as_text(v) or "").lower()
    return SOURCE_TO_ENUM.get(t, "DIRECT_OTHER")

NAICS_RE = re.compile(r"\((\d[\d\-]*)\)")
def parse_naics_and_industry(industry, mfg, ptype):
    # most specific description available
    desc = as_text(industry) or as_text(mfg) or as_text(ptype)
    # most specific (longest all-digit) code across the three columns
    best = None
    for cell in (industry, mfg, ptype):
        for m in NAICS_RE.finditer(str(cell or "")):
            code = m.group(1)
            if code.isdigit() and (best is None or len(code) > len(best)):
                best = code
    # strip the "(code)" from the description for cleanliness
    if desc:
        desc = re.sub(r"\s*\([\d\-]+\)\s*$", "", desc).strip() or None
    return best, desc

EMAIL_RE = re.compile(r"[\w.+-]+@[\w-]+\.[\w.-]+")
def parse_contact(v):
    t = as_text(v)
    if t is None: return None, None
    email = None
    m = EMAIL_RE.search(t)
    if m: email = m.group(0)
    name = t.split("\n")[0].strip() or None
    if name and EMAIL_RE.fullmatch(name):  # first line was the email itself
        name = None
    return name, email

def utilities_for(row, idx):
    out = []
    for field, utype in (("power", "ELECTRICITY"), ("water", "WATER"),
                         ("sewer", "WASTEWATER"), ("gas", "GAS")):
        raw = util_raw(get(row, idx, field))
        if not raw:
            continue
        p = parse_utility(utype, raw)
        out.append({
            "type": utype,
            "rawValue": raw,
            "normalizedValue": p["normalizedValue"],
            "normalizedUnit": p["normalizedUnit"],
            "flagged": p["flagged"],
            "assumptionNote": p["assumptionNote"],
            "datapoints": p["datapoints"],
        })
    return out

def job_phases(row, idx):
    n = as_int(get(row, idx, "jobs"))
    if n is None: return []
    return [{"count": n, "timeframe": "Total (from FY26 tracker)"}]

def building_needs(row, idx):
    sf_min = as_num(get(row, idx, "sf_min")); sf_max = as_num(get(row, idx, "sf_max"))
    ac_min = as_num(get(row, idx, "ac_min")); ac_max = as_num(get(row, idx, "ac_max"))
    bits = []
    if sf_min is not None or sf_max is not None:
        if sf_min is not None and sf_max is not None and sf_max != sf_min:
            bits.append(f"Building {int(sf_min):,}\u2013{int(sf_max):,} SF")
        else:
            bits.append(f"Building {int(sf_min or sf_max):,} SF")
    if ac_min is not None or ac_max is not None:
        if ac_min is not None and ac_max is not None and ac_max != ac_min:
            bits.append(f"{_n(ac_min)}\u2013{_n(ac_max)} acres")
        else:
            bits.append(f"{_n(ac_min if ac_min is not None else ac_max)} acres")
    text = "; ".join(bits) + " (range from FY26 tracker)" if bits else None
    return sf_min, ac_min, text

def _n(x):
    return str(int(x)) if x == int(x) else str(x)

def qual_notes(row, idx):
    out = []
    cons = as_text(get(row, idx, "consultant"))
    if cons: out.append({"label": "Company / Site Selector (real identity)", "content": cons})
    comp = as_text(get(row, idx, "competing"))
    if comp: out.append({"label": "Competing States", "content": comp})
    bb = as_text(get(row, idx, "bluebonnet")); pn = as_text(get(row, idx, "pedernales"))
    if bb or pn:
        parts = []
        if bb: parts.append(f"Bluebonnet co-op sites: {bb}")
        if pn: parts.append(f"Pedernales co-op sites: {pn}")
        out.append({"label": "Sites Offered", "content": "\n".join(parts)})
    return out

def narrative(row, idx):
    notes = as_text(get(row, idx, "notes")); upd = as_text(get(row, idx, "updates"))
    parts = []
    if notes: parts.append(notes)
    if upd: parts.append("UPDATES:\n" + upd)
    return "\n\n".join(parts) if parts else None

def transportation(row, idx):
    port = as_text(get(row, idx, "port"))
    if port and port.lower() not in ("no", "n/a", "-"):
        return f"Port: {port}"
    return None

# --- per-row project builder ----------------------------------------------
def build_project(row, idx, *, force_stage=None, archived=False):
    name = as_text(get(row, idx, "name"))
    if not name: return None
    submitted = as_date(get(row, idx, "submitted"))
    status = as_text(get(row, idx, "status"))
    stage = force_stage or stage_from_status(status, submitted is not None)
    naics, industry_desc = parse_naics_and_industry(
        get(row, idx, "industry"), get(row, idx, "mfg_type"), get(row, idx, "proj_type"))
    cname, cemail = parse_contact(get(row, idx, "contact"))
    sf_min, ac_min, needs = building_needs(row, idx)
    # projectType: only the stray "Expansion" carries a real project-type meaning
    ptype = "Expansion" if (status or "").strip().lower() == "expansion" else None
    proj = {
        "codename": name,
        "stage": stage,
        "leadSource": lead_source(get(row, idx, "source")),
        "sourceContactName": cname,
        "sourceContactEmail": cemail,
        "companyLocationRaw": as_text(get(row, idx, "origin")),
        "naicsCode": naics,
        "industryDescription": industry_desc,
        "projectType": ptype,
        "capexTotal": as_num(get(row, idx, "investment")),
        "minBuildingSqFt": sf_min,
        "minAcreage": ac_min,
        "buildingSizeNeeds": needs,
        "railPreference": pref(get(row, idx, "rail")),
        "existingBuildingPreference": pref(get(row, idx, "existing")),
        "transportationNotes": transportation(row, idx),
        "rfiReceivedDate": as_date(get(row, idx, "active")),
        "responseSubmittedDate": submitted,
        "narrative": narrative(row, idx),
        "noSubmissionReason": as_text(get(row, idx, "no_sub_why")),
        "archived": archived,
        "jobPhases": job_phases(row, idx),
        "utilities": utilities_for(row, idx),
        "qualitativeNotes": qual_notes(row, idx),
        "_sourceStatus": status,  # kept for the audit summary only
    }
    return proj

# --- main ------------------------------------------------------------------
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

projects = []
def ingest(sheet, **kw):
    ws = wb[sheet]; idx = build_index(ws)
    got = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        p = build_project(row, idx, **kw)
        if p:
            p["_sheet"] = sheet
            projects.append(p); got += 1
    return got

n1 = ingest("FY26 Leads")
n2 = ingest("No Submission", force_stage="NO_SUBMISSION")
n3 = ingest("FY25 Follow-Ups", archived=True)

# Site tours -> visits (attached later by codename match)
ws = wb["Site Tours"]; tidx = {hnorm(h): i for i, h in enumerate(
    next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))}
def tget(row, key):
    i = tidx.get(key)
    return row[i] if (i is not None and i < len(row)) else None
visits = []
for row in ws.iter_rows(min_row=2, values_only=True):
    nm = as_text(tget(row, "project name"))
    dt = as_date(tget(row, "tour date"))
    if not nm or not dt: continue
    vtype = as_text(tget(row, "visit type"))
    sites = as_text(tget(row, "site name(s)"))
    note_bits = []
    if vtype: note_bits.append(vtype)
    if sites: note_bits.append("Sites: " + sites.replace("\n", ", "))
    visits.append({"codename": nm, "visitDate": dt,
                   "note": " \u2014 ".join(note_bits) if note_bits else None})

out = {"projects": projects, "visits": visits}
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(out, f, ensure_ascii=False, indent=2)

# --- summary ---------------------------------------------------------------
from collections import Counter
print(f"Projects: {len(projects)}  (FY26 Leads {n1} / No Submission {n2} / FY25 {n3})")
print("Stages:", dict(Counter(p["stage"] for p in projects)))
print("Lead sources:", dict(Counter(p["leadSource"] for p in projects)))
codenames = set(p["codename"] for p in projects)
matched = sum(1 for v in visits if v["codename"] in codenames)
print(f"Site visits: {len(visits)}  (match a project: {matched}, unmatched: {len(visits)-matched})")
print("Unmatched tour projects:", sorted({v['codename'] for v in visits if v['codename'] not in codenames}))
# data-quality flags
missing_active = [p["codename"] for p in projects if not p["rfiReceivedDate"]]
print("Projects missing Active(RFI) date:", len(missing_active), missing_active[:10])
print("Projects with capex:", sum(1 for p in projects if p["capexTotal"] is not None))
print("Projects with >=1 utility:", sum(1 for p in projects if p["utilities"]))
print("Projects with Sites Offered note:", sum(1 for p in projects if any(q["label"]=="Sites Offered" for q in p["qualitativeNotes"])))
wb.close()
print(f"\nWrote {OUT}")
